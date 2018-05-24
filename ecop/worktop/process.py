import base64
import json
from decimal import Decimal
from datetime import datetime, timedelta
from dateutil import tz

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment
from pyramid_rpc.jsonrpc import jsonrpc_method
from pyramid.view import view_config
from pyramid.response import Response

from hm.lib.number import decimal_round
from weblibs.camunda import camundaClient as cc
from weblibs.jsonrpc import RPCUserError, parseDate
from webmodel.consts import ORDER_SOURCE, SPECIAL_PARTY
from webmodel.order import SalesOrder, PurchaseOrder
from webmodel.item import Item
from webmodel.validator import isOrderId, isMobile

from ecop.base import RpcBase, DocBase
from ecop.worktop.validator import isIkeaOrderId
from ecop.worktop.utils import addItemInfo

"""
In case a process instance needs manual correction, use:

POST /process-instance/abe0b544-33f3-11e8-8a21-0242ac110005/modification
{
  "instructions": [{
    "type": "startBeforeActivity",
    "activityId": "ConfirmMeasurementDate"
  },{
    "type": "cancel",
    "activityId": "TakeMeasurement"
  }]
}
"""

__StatusNames__ = {
    'EXTERNALLY_TERMINATED': '已取消',
    'COMPLETED': '已完成',
    'INTERNALLY_TERMINATED': '已取消'
}

tzLocal = tz.tzlocal()


def searchProcess(cond, request, countOnly=False, maxRows=50):
    """ Note the startDate and endDate will be passed in UTC """
    cond['storeId'] = request.user.extraData['worktop'].get('storeId')

    params = {
        'processDefinitionKey': 'worktop',
        'sorting': [{
            'sortBy': 'startTime',
            'sortOrder': 'desc'
        }]
    }

    searchText = cond.get('searchText')
    showCompleted = cond.get('completed')

    if searchText:
        if isOrderId(searchText):
            params['processInstanceBusinessKey'] = searchText
        elif isIkeaOrderId(searchText):
            params['variables'] = [{
                'name': 'externalOrderId',
                'operator': 'eq',
                'value': searchText.upper()
            }]
        elif isMobile(searchText):
            params['variables'] = [{
                'name': 'customerMobile',
                'operator': 'eq',
                'value': searchText
            }]
        else:
            params['variables'] = [{
                'name': 'customerName',
                'operator': 'eq',
                'value': searchText
            }]
    else:
        parseDate(cond, fields=['startDate', 'endDate'])
        startDate, endDate = cond['startDate'], cond['endDate']
        endDate = endDate + timedelta(1)

        if (endDate - startDate) > timedelta(365):
            raise RPCUserError('订单查询时间跨度不能大于１年。')

        params['startedBefore'] = endDate
        params['startedAfter'] = startDate
        if showCompleted:
            params['finished'] = 'true'

    storeId = cond['storeId']
    if storeId:
        variables = params.setdefault('variables', [])
        variables.append({
            'name': 'storeId',
            'operator': 'eq',
            'value': storeId
        })

    if countOnly:
        ret = cc.makeRequest(
            '/history/process-instance/count', 'post',
            params,
        )
        if ret['count'] > 500:
            raise RPCUserError('单次导出结果大于500条，请搜索条件再')
        return ret['count']
    else:
        ret = cc.makeRequest(
            '/history/process-instance', 'post',
            params,
            urlParams={'maxResults': maxRows},
            withProcessVariables=(
                'externalOrderId', 'customerName', 'storeId', 'orderItems',
                'shippingDate', 'receivingDate',
                'actualMeasurementDate', 'confirmedMeasurementDate',
                'scheduledMeasurementDate', 'actualInstallationDate',
                'confirmedInstallationDate', 'scheduledInstallationDate'
            ),
            processInstanceIdField='id', hoistProcessVariables=True
        )

        # this filters out CANCELED processes
        if not searchText and showCompleted:
            ret = [p for p in ret if p['state'] == 'COMPLETED']

        #
        # Prepare for display by adding additional infos:
        #  * add model to orderItems as only itemId is stored
        #  * add human readable status text
        #
        currentTime = datetime.now()
        for p in ret:
            # The instance variables of Date type are parsed correctly, but the
            # process property is not. We will do the parse here.
            parseDate(p, fields=['startTime'])
            p['startTime'] = p['startTime'].astimezone(
                tzLocal).replace(tzinfo=None)

            # calculate the duration of the process.
            if 'TERMINATED' not in p['state']:
                start = p.get('actualMeasurementDate') or \
                    p.get('scheduledMeasurementDate') or p['startTime']
                end = p.get('actualInstallationDate') or currentTime
                delta = end - start
                if delta.total_seconds() > 0:
                    p['duration'] = delta.days + decimal_round(
                        Decimal(delta.seconds / 86400), Decimal('0.1'))

            state = p.pop('state')
            if state == 'ACTIVE':
                if p.get('actualInstallationDate'):
                    text = '已安装'
                elif p.get('confirmedInstallationDate'):
                    text = '待安装'
                elif p.get('receivingDate'):
                    text = '已收货'
                elif p.get('shippingDate'):
                    text = '已发货'
                elif p.get('actualMeasurementDate') or \
                        not p.get('scheduledMeasurementDate'):
                    text = '生产中'
                elif p.get('confirmedMeasurementDate') or \
                        p.get('scheduledMeasurementDate'):
                    text = '待测量'
                else:
                    text = '进行中'
                p['statusText'] = text
            else:
                p['statusText'] = __StatusNames__[state]

            ois = p.get('orderItems')
            if ois:
                addItemInfo(ois)

        return ret


class ProcessJSON(RpcBase):
    @jsonrpc_method(endpoint='rpc', method='bpmn.process.start')
    def startProcess(self, params):
        params = parseDate(
            params,
            fields=['scheduledMeasurementDate', 'scheduledInstallationDate']
        )

        if not params['scheduledInstallationDate']:
            params.pop('scheduledInstallationDate')

        if not params['isMeasurementRequested']:
            params.pop('scheduledMeasurementDate', None)
            params['productionDrawing'] = params['orderFile']

        externalOrderId = params.get('externalOrderId')
        if externalOrderId and \
                self.sess.query(SalesOrder).filter_by(
                    orderSource=ORDER_SOURCE.IKEA,
                    externalOrderId=externalOrderId
                ).all():
            raise RPCUserError('该订单号已存在，不能重复提交')

        order = SalesOrder(
            customerId=int(SPECIAL_PARTY.BE),
            creatorId=self.request.user.partyId,
            regionCode=params['customerRegionCode'],
            streetAddress=params['customerStreet'],
            recipientName=params['customerName'],
            recipientMobile=params['customerMobile'],
            orderSource=ORDER_SOURCE.IKEA
        )

        if params['isMeasurementRequested']:
            itemMeasure = self.sess.query(Item).get(10023928)
            order.addItem(item=itemMeasure, quantity=1)
        itemInstall = self.sess.query(Item).get(10023929)
        order.addItem(item=itemInstall, quantity=1)

        self.sess.add(order)
        self.sess.flush()

        po = PurchaseOrder(
            supplierId=10025188,
            customerId=int(SPECIAL_PARTY.BE),
            creatorId=self.request.user.partyId,
            regionCode=params['customerRegionCode'],
            streetAddress=params['customerStreet'],
            recipientName=params['customerName'],
            recipientMobile=params['customerMobile'],
            relatedOrderId=order.orderId
        )
        if params['isMeasurementRequested']:
            po.addItem(item=itemMeasure, quantity=1)
        po.addItem(item=itemInstall, quantity=1)
        self.sess.add(po)

        if externalOrderId:
            order.externalOrderId = externalOrderId
        else:
            params['externalOrderId'] = str(order.orderId)

        # add orderId also as a process instance variable
        params['orderId'] = order.orderId

        cc.makeRequest(
            f'/process-definition/key/worktop/start', 'post',
            {'businessKey': order.orderId},
            variables=params
        )

    @jsonrpc_method(endpoint='rpc', method='bpmn.process.list')
    def getProcessList(self, cond):
        """
        Search process history with the given conditions.

        :param kwargs: a dictionary with keywords `searchText`, `startDate`,
            `endDate` or `completed`.
            `searchText` could mean orderId, customerMobile or customerName.
            The `startDate` and `endDate` refer to the process start date.
            If `completed` is set to true, only completed orders will be
            returned.
            Note if `searchText` is non-empty, the other fields are ignored.
        """
        return searchProcess(cond, self.request)

    @jsonrpc_method(endpoint='rpc', method='bpmn.process.list.checkCount')
    def getProcessListCount(self, cond):
        """
        For process list export to Excel we don't limit the returned number of
        rows to 50, but to 500.

        This is a pre-check on the would-be number of returns. It throws
        RPCUserError if the above condition is not satisfied.
        """
        return searchProcess(cond, self.request, countOnly=True)

    @jsonrpc_method(endpoint='rpc', method='bpmn.variable.get')
    def getProcessVariables(self, processInstanceId):
        variables = cc.parseVariables(
            cc.makeRequest(f'/history/variable-instance', 'post', {
                'processInstanceIdIn': [processInstanceId]
            }, urlParams={'deserializeValues': 'false'}))

        ois = variables.get('orderItems')
        if ois:
            addItemInfo(ois)

        return variables


@view_config(route_name='processlist')
class ProcessList(DocBase):
    """
    Export a process list to Excel
    """

    def __call__(self):
        cond = json.loads(base64.b64decode(self.request.params['c']))
        cond['download'] = True

        processes = searchProcess(cond, self.request, maxRows=500)

        alCenter = Alignment(horizontal='center')
        alLeft = Alignment(horizontal='left')
        alWrap = Alignment(wrapText=True)

        wb = Workbook()
        ws = wb.active

        for (idx, t) in enumerate((
            '订单号', '商场号', '顾客姓名', '台面', '发起时间', '预约测量日',
            '确认测量日', '实际测量日', '收货日期', '预约安装日', '确认安装日',
                '实际安装日', '耗时', '状态')):
            cell = ws[f'{chr(ord("A")+idx)}1']
            cell.value = t
            cell.alignment = alCenter

        for (row, p) in enumerate(processes):
            ws[f'A{row+2}'] = p['externalOrderId']
            ws[f'B{row+2}'] = p['storeId']
            ws[f'C{row+2}'] = p['customerName']

            ois = p.get('orderItems')
            if ois:
                ws[f'D{row+2}'] = '\n'.join(
                    [f'{oi["model"]}*{oi["quantity"]}' for oi in ois])

            # openpyxl can not handle timezone properly
            ws[f'E{row+2}'] = p['startTime']
            ws[f'F{row+2}'] = p.get('scheduledMeasurementDate')
            ws[f'G{row+2}'] = p.get('confirmedMeasurementDate')
            ws[f'H{row+2}'] = p.get('actualMeasurementDate')
            ws[f'I{row+2}'] = p.get('receivingDate')
            ws[f'J{row+2}'] = p.get('scheduledInstallationDate')
            ws[f'K{row+2}'] = p.get('confirmedInstallationDate')
            ws[f'L{row+2}'] = p.get('actualInstallationDate')
            ws[f'M{row+2}'] = p.get('duration')
            ws[f'N{row+2}'] = p['statusText']

            ws[f'D{row+2}'].alignment = alWrap
            ws[f'E{row+2}'].alignment = alLeft
            for col in 'FGHIJKL':
                ws[f'{col}{row+2}'].number_format = 'yyyy-mm-dd'
                ws[f'{col}{row+2}'].alignment = alLeft

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        for col in 'FGHIJKL':
            ws.column_dimensions[col].width = 12

        body = save_virtual_workbook(wb)
        fileName = cond['startDate'].astimezone(tzLocal).strftime('%Y%m%d') \
            + '_' + cond['endDate'].astimezone(tzLocal).strftime('%Y%m%d')
        response = Response(
            content_type='application/octet-stream',
            content_disposition=f'attachment; filename="{fileName}.xlsx"',
            content_length=len(body),
            body=body)
        return response
